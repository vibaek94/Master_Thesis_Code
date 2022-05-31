import openpyxl as openpyxl
import pandas as pd
import time
import keras
import numpy as np
from itertools import islice
from numpy import array
from numpy import hstack
import tensorflow as tf
from numpy import mean
from numpy import std
import matplotlib.pyplot as plt
from matplotlib import cm
from matplotlib.colors import ListedColormap, LinearSegmentedColormap
# split a multivariate sequence into samples
# importing os module
import os
import seaborn as sb
from sklearn.metrics import mean_squared_error
from sklearn.metrics import mean_absolute_percentage_error
from sklearn.metrics import r2_score
def load_data(path,n_steps = 10):

    #used to load the dataset from the path given and splits the data into a test set of 10 subjects, a training set of 19 subjects.
    # and a validation set of 4 subjects. The activity label of the is also extracted, and the subject ID.

    wb = openpyxl.load_workbook(path, read_only=True)
    sheet = wb.worksheets[0]
    ws = wb.active

    data = ws.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)

    df1 = pd.DataFrame(data, index=idx, columns=cols)
    df1['ID'] = df1.index
    REE,Weight,id,correction = list(),list(),list(),list()
    for x in range(df1['ID'].min(), df1['ID'].max() + 1, 1):
        if (x in df1['ID']):
            temp = df1.loc[lambda df1: df1['ID'] == x]
            temp1 = temp.loc[lambda temp: temp['ActivityType'] < 4]
            corr = temp['vo2']-((temp['vo2n'])*50*temp1['Weight'].mean()*100+temp1['vo2'].mean())
            correction.append(corr.mean())
            REE.append(temp1['vo2'].mean())
            Weight.append(temp1['Weight'].mean()*100)
            id.append(x)
    feat = pd.DataFrame(list(zip(REE, Weight,correction)),
                      columns=['Resting EE', 'Weight','Correction'],index=id)
    # 10 sec
    # defines the number of samples in the combined training and validation set, corresponds to 23 subjects
    train_size = 9034
    train, test = df1.values[0:train_size, ], df1.values[train_size:len(df1.values), ]
    # defines the number of samples in the training set corresponds to 19 subjects
    val_size = 7498
    train, val = train[0:val_size, :], train[val_size:len(train)]

    # choose a number of time steps
    # convert into input/output

    X_train, y_train, VO2_train, id_train = preparedata(train, n_steps=n_steps)
    X_val, y_val, VO2_val, id_val = preparedata(val, n_steps)


    X_test, y_test, VO2_test,id_test = preparedata(test, n_steps=n_steps)
    # select the column in the test set that contains the activity labels
    # calls a function that takes the median activity label of the n_steps samples
    activity_data = test[:, 4]
    # calls a function that takes the median activity label of the n_steps samples
    act_label = activity_label(activity_data, n_steps)
    # calls the same function but with subject id instead
    subject_id_test = activity_label(idx[train_size:len(idx)],n_steps)
    # calls the same function but with subject id instead

    return X_train, y_train, X_val , y_val, X_test , y_test, feat, act_label, VO2_test, id_test

def activity_label(data,n_steps):
    # assigns the median value for a feature in n_steps samples
    act_label = list()
    for i in range(len(data)):
        end_ix = i + n_steps
        act_slice=data[i:end_ix]
        median_act=np.median(act_slice)

        if end_ix > len(data):
            break
        # the median defaults to the prior value when the median function returns a value that is not present in the dataset
        if median_act not in act_slice:
            index = int(n_steps/2-1)
            median_act = act_slice[index]
        act_label.append(median_act)
    return act_label

def calc_B2_VO2(feat,prediction,y_test,VO2_test,id_test):
    df = pd.DataFrame(list(zip(prediction, y_test,id_test,VO2_test)),
                      columns=['predicted_VO2n', 'actual_y','subject_ID','VO2'])
    predicted_VO2 = pd.Series()
    for x in range(28,38):
        temp = df.loc[lambda df: df['subject_ID'] == x]
        temp_pred_VO2 = temp['predicted_VO2n']
        temp_pred_VO2 = temp_pred_VO2 * 50
        temp_pred_VO2 = temp_pred_VO2 * feat.loc[x][1]
        temp_pred_VO2 = temp_pred_VO2 + feat.loc[x][0]
        temp_pred_VO2 = temp_pred_VO2 + feat.loc[x][2]
        predicted_VO2 = predicted_VO2.append(temp_pred_VO2,ignore_index=True)
    return predicted_VO2

def split_sequence(sequences, n_steps):
    X, y, VO2, id = list(), list(), list(),list()
    for i in range(len(sequences)):
        # find the end of this pattern
        end_ix = i + n_steps
        # check if we are beyond the dataset
        if end_ix > len(sequences):
            break
        # gather input and output parts of the pattern
        seq_x,seq_VO2,seq_id, seq_y = sequences[i:end_ix, :-3],sequences[end_ix - 1, -3],sequences[end_ix - 1, -2], sequences[end_ix - 1, -1]
        X.append(seq_x)
        y.append(seq_y)
        VO2.append(seq_VO2)
        id.append(seq_id)

    return array(X), array(y), array(VO2),array(id)
#Preparing the data for multivariate LSTM
def preparedata(data, n_steps):
    # define input sequence
    #in_seq1 = data[:, 7]
    #1 is AGI
    #7 is AGI_weight
    #2 is MAD
    #vo2,incl,agi,mad,activitytype,gender,weight,height,leg,arm,waist,vo2n,vo2r
    #0  ,   1,  2,  3,           4,     5,     6,     7,  8,  9,   10,  11,  12

    in_seq1 = data[:, 2]
    in_seq2 = data[:, 1]
    in_seq3 = data[:, 7]
    in_seq4 = data[:, 8]
    in_seq5 = data[:, 9]
    VO2_seq = data[:,0]
    id_seq = data[:,13]
    out_seq = data[:, 11]
    # convert to [rows, columns] structure
    in_seq1 = in_seq1.reshape((len(in_seq1), 1))
    in_seq2 = in_seq2.reshape((len(in_seq2), 1))
    in_seq3 = in_seq3.reshape((len(in_seq3), 1))
    in_seq4 = in_seq4.reshape((len(in_seq4), 1))
    in_seq5 = in_seq5.reshape((len(in_seq5), 1))
    VO2_seq = VO2_seq.reshape((len(VO2_seq), 1))
    id_seq = id_seq.reshape((len(id_seq), 1))
    out_seq = out_seq.reshape((len(out_seq), 1))
    # horizontally stack columns
    dataset = hstack((in_seq1, in_seq2, in_seq3, in_seq4,in_seq5,VO2_seq ,id_seq, out_seq))
    #dataset = hstack((in_seq1, out_seq))
    # choose a number of time steps
    #n_steps = 5
    # convert into input/output
    X, y, VO2, id = split_sequence(dataset, n_steps)
    return X, y, VO2, id

def load_data_MAD(path,n_steps = 10):
    # used to load the dataset from the path given and splits the data into a test set of 10 subjects, a training set of 19 subjects.
    # and a validation set of 4 subjects. The activity label of the is also extracted, and the subject ID.

    wb = openpyxl.load_workbook(path, read_only=True)
    sheet = wb.worksheets[0]
    ws = wb.active

    data = ws.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)

    df1 = pd.DataFrame(data, index=idx, columns=cols)
    df1['ID'] = df1.index
    REE, Weight, id, correction = list(), list(), list(), list()
    for x in range(df1['ID'].min(), df1['ID'].max() + 1, 1):
        if (x in df1['ID']):
            temp = df1.loc[lambda df1: df1['ID'] == x]
            temp1 = temp.loc[lambda temp: temp['ActivityType'] < 4]
            corr = temp['vo2'] - ((temp['vo2n']) * 50 * temp1['Weight'].mean() * 100 + temp1['vo2'].mean())
            correction.append(corr.mean())
            REE.append(temp1['vo2'].mean())
            Weight.append(temp1['Weight'].mean() * 100)
            id.append(x)
    feat = pd.DataFrame(list(zip(REE, Weight, correction)),
                        columns=['Resting EE', 'Weight', 'Correction'], index=id)
    # 10 sec
    # defines the number of samples in the combined training and validation set, corresponds to 23 subjects
    train_size = 9034
    train, test = df1.values[0:train_size, ], df1.values[train_size:len(df1.values), ]
    # defines the number of samples in the training set corresponds to 19 subjects
    val_size = 7498
    train, val = train[0:val_size, :], train[val_size:len(train)]

    # choose a number of time steps
    # convert into input/output

    X_train, y_train, VO2_train, id_train = preparedata_MAD(train, n_steps=n_steps)
    X_val, y_val, VO2_val, id_val = preparedata_MAD(val, n_steps)

    X_test, y_test, VO2_test, id_test = preparedata_MAD(test, n_steps=n_steps)
    # select the column in the test set that contains the activity labels
    # calls a function that takes the median activity label of the n_steps samples
    activity_data = test[:, 4]
    # calls a function that takes the median activity label of the n_steps samples
    act_label = activity_label(activity_data, n_steps)
    # calls the same function but with subject id instead

    return X_train, y_train, X_val, y_val, X_test, y_test, feat, act_label, VO2_test, id_test

def preparedata_MAD(data, n_steps):
    # define input sequence
    #in_seq1 = data[:, 7]
    #1 is AGI
    #7 is AGI_weight
    #2 is MAD
    #vo2,incl,agi,mad,activitytype,gender,weight,height,leg,arm,waist,vo2n,vo2r
    #0  ,   1,  2,  3,           4,     5,     6,     7,  8,  9,   10,  11,  12

    in_seq1 = data[:, 3]
    in_seq2 = data[:, 1]
    in_seq3 = data[:, 7]
    in_seq4 = data[:, 8]
    in_seq5 = data[:, 9]
    VO2_seq = data[:,0]
    id_seq = data[:,13]
    out_seq = data[:, 11]
    # convert to [rows, columns] structure
    in_seq1 = in_seq1.reshape((len(in_seq1), 1))
    in_seq2 = in_seq2.reshape((len(in_seq2), 1))
    in_seq3 = in_seq3.reshape((len(in_seq3), 1))
    in_seq4 = in_seq4.reshape((len(in_seq4), 1))
    in_seq5 = in_seq5.reshape((len(in_seq5), 1))
    VO2_seq = VO2_seq.reshape((len(VO2_seq), 1))
    id_seq = id_seq.reshape((len(id_seq), 1))
    out_seq = out_seq.reshape((len(out_seq), 1))
    # horizontally stack columns
    dataset = hstack((in_seq1, in_seq2, in_seq3, in_seq4,in_seq5,VO2_seq ,id_seq, out_seq))
    #dataset = hstack((in_seq1, out_seq))
    # choose a number of time steps
    #n_steps = 5
    # convert into input/output
    X, y, VO2, id = split_sequence(dataset, n_steps)
    return X, y, VO2, id

def load_data_VO2(path,n_steps = 10):

    #used to load the dataset from the path given and splits the data into a test set of 10 subjects, a training set of 19 subjects.
    # and a validation set of 4 subjects. The activity label of the is also extracted, and the subject ID.

    wb = openpyxl.load_workbook(path, read_only=True)
    sheet = wb.worksheets[0]
    ws = wb.active

    data = ws.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)

    df1 = pd.DataFrame(data, index=idx, columns=cols)

    # 10 sec
    # defines the number of samples in the combined training and validation set, corresponds to 23 subjects
    train_size = 9034
    train, test = df1.values[0:train_size, ], df1.values[train_size:len(df1.values), ]
    # defines the number of samples in the training set corresponds to 19 subjects
    val_size = 7498
    train, val = train[0:val_size, :], train[val_size:len(train)]

    # choose a number of time steps
    # convert into input/output

    X_train, y_train = preparedata_VO2(train, n_steps=n_steps)
    X_val, y_val = preparedata_VO2(val, n_steps)


    X_test, y_test = preparedata_VO2(test, n_steps=n_steps)
    # select the column in the test set that contains the activity labels
    activity_data = test[:, 4]
    # calls a function that takes the median activity label of the n_steps samples
    act_label = activity_label(activity_data, n_steps)
    # calls the same function but with subject id instead
    subject_id_test = activity_label(idx[train_size:len(idx)],n_steps)

    return X_train, y_train , X_val , y_val , X_test , y_test , act_label, subject_id_test

def preparedata_VO2(data, n_steps):
    # define input sequence
    #in_seq1 = data[:, 7]
    #1 is AGI
    #7 is AGI_weight
    #2 is MAD
    #vo2,incl,agi,mad,activitytype,gender,weight,height,leg,arm,waist,vo2n,vo2r
    #0  ,   1,  2,  3,           4,     5,     6,     7,  8,  9,   10,  11,  12

    in_seq1 = data[:, 2]
    in_seq2 = data[:, 1]
    in_seq3 = data[:, 7]
    in_seq4 = data[:, 8]
    in_seq5 = data[:, 9]
    out_seq = data[:, 0]
    # convert to [rows, columns] structure
    in_seq1 = in_seq1.reshape((len(in_seq1), 1))
    in_seq2 = in_seq2.reshape((len(in_seq2), 1))
    in_seq3 = in_seq3.reshape((len(in_seq3), 1))
    in_seq4 = in_seq4.reshape((len(in_seq4), 1))
    in_seq5 = in_seq5.reshape((len(in_seq5), 1))
    out_seq = out_seq.reshape((len(out_seq), 1))
    # horizontally stack columns
    dataset = hstack((in_seq1, in_seq2, in_seq3, in_seq4, in_seq5, out_seq))
    #dataset = hstack((in_seq1, out_seq))
    # choose a number of time steps
    #n_steps = 5
    # convert into input/output
    X, y = split_sequence(dataset, n_steps)
    return X, y



def save_model(model_name, model):
    # saves a model in the correct directory
    # Parent Directory path
    parent_dir = r'C:\Users\jcb\Desktop\speciale\models'
    # Path
    path = os.path.join(parent_dir, model_name)
    model.save(path + '\\' + model_name + '.h5')
def bland_altman_plot(data1, data2, *args, **kwargs):
    data1     = np.asarray(data1)
    data2     = np.asarray(data2)
    mean      = np.mean([data1, data2], axis=0)
    diff      = data1 - data2                   # Difference between data1 and data2
    md        = np.mean(diff)                   # Mean of the difference
    sd        = np.std(diff, axis=0)            # Standard deviation of the difference

    plt.scatter(mean, diff, *args, **kwargs)
    plt.axhline(md,           color='gray', linestyle='--')
    plt.axhline(md + 1.96*sd, color='gray', linestyle='--')
    plt.axhline(md - 1.96*sd, color='gray', linestyle='--')


def plot_model(model_name,pred_y,test_y,activity_label,fit_history=None,subject_ID=None):
    # creates a dataframe containing predicted y values, actual y values, activity type and subject ID
    df = pd.DataFrame(list(zip(pred_y, test_y,activity_label,subject_ID.astype(int))),
                      columns=['predicted_y', 'actual_y','activity_label','subject_ID'])

    # Parent Directory path
    parent_dir = r'C:\Users\jcb\Desktop\speciale\models'
    # Path
    path = os.path.join(parent_dir, model_name)
    #makes directory for the model name
    os.mkdir(path)

    # plot of the values for predicted y and actual y. also plots R^2, MSE and MAPE
    mse = mean_squared_error(test_y, pred_y)
    mape = mean_absolute_percentage_error(test_y, pred_y)
    print('MSE: %.3f MAPE: %.3f%%' % (mse, mape))
    plt.style.use('ggplot')
    plt.figure(11)
    plt.plot(pred_y, label="Predicted y")
    plt.plot(test_y, label="Actual y")
    plt.legend(['R^2: %.3f' % r2_score(test_y, pred_y),'MSE: %.3f MAPE: %.3f%%' % (mse, mape)],loc='lower right')
    print(path)
    plt.savefig(path+'\\'+model_name+'_fit.png')
    plt.close()
    #plt.show()

    # scatterplot of pred_y related to test_y
    plt.figure(12, figsize=(10, 10))
    plt.scatter(test_y, pred_y, c='crimson')
    p1 = max(max(pred_y), max(test_y))
    p2 = min(min(pred_y), min(test_y))
    plt.plot([p1, p2], [p1, p2], 'b-')
    plt.xlabel('True Values', fontsize=15)
    plt.ylabel('Predictions', fontsize=15)
    plt.axis('equal')
    m, b = np.polyfit(test_y, pred_y, 1)
    plt.plot(test_y, m * test_y + b)
    plt.savefig(path+'\\'+model_name+'_scatter.png')
    plt.close()
    #plt.show()

    #plot of MSE and Val_MSE from model_fit per Epoch
    if fit_history != None:
        plt.figure(13, figsize=(10, 10))
        plt.plot(fit_history.history['mean_squared_error'])

        plt.plot(fit_history.history['val_mean_squared_error'])
        plt.title('model MSE')
        plt.xlabel('Epochs', fontsize=15)
        plt.ylabel('MSE', fontsize=15)

        plt.legend(['train', 'validation'], loc='upper left')
        plt.savefig(path+'\\'+model_name+'_history.png')
        plt.close()
        #plt.show()

    # scatterplot with activity labels
    activity_plot=sb.scatterplot(data=df, x="actual_y", y="predicted_y", hue="activity_label",palette="deep")
    x0, x1 = activity_plot.get_xlim()
    y0, y1 = activity_plot.get_ylim()
    lims = [max(x0, y0), min(x1, y1)]
    activity_plot.plot(lims, lims, '-r')

    fig = activity_plot.get_figure()
    fig.savefig(path+'\\'+model_name+'_activity_scatter.png')
    fig.clear()
    id_plot(df,path,model_name)


def id_plot(df,path,model_name):
    for x in range(df['subject_ID'].min(), df['subject_ID'].max() + 1, 1):
        if (x in df['subject_ID']):
            subject1 = df.loc[df['subject_ID'] == x]
            mse = mean_squared_error(subject1['predicted_y'], subject1['actual_y'])
            mape = mean_absolute_percentage_error(subject1['predicted_y'], subject1['actual_y'])
            print('MSE: %.3f MAPE: %.3f%%' % (mse, mape))
            print('MSE: %.3f MAPE: %.3f%%' % (mse, mape))
            plt.style.use('ggplot')
            plt.figure(11)
            plt.plot(range(len(subject1['predicted_y'])),subject1['predicted_y'], label="Predicted y")
            plt.plot(range(len(subject1['predicted_y'])),subject1['actual_y'], label="Actual y")
            plt.legend(['R^2: %.3f' % r2_score(test_y, pred_y), 'MSE: %.3f MAPE: %.3f%%' % (mse, mape)],
                       loc='lower right')
            #plt.legend(loc='upper left')
            plt.xlabel('Samples')
            plt.ylabel('VO2')

            plt.savefig(path + '\\' + model_name + '_fit_subject_{}.png'.format(x))
            plt.close()



def model_load(model_name):
    # loads a model based on the model name
    parent_dir = r'C:\Users\jcb\Desktop\speciale\models'
    path = parent_dir +'\\'+ model_name + '\\'+model_name
    loaded_model = keras.models.load_model(path+'.h5')
    return loaded_model

def shift_act_labels(act_label):
    # shifts the activity labels to remove missing values as activity_type = 11.
    # Also provides a list with the names corresponding to the values in activity_type
    act_name = ['sitting', 'sitting, playing tablet', 'standing, playing tablet', 'walking (preferred speed)',
                'walking (brisk speed)', 'running', 'basket', 'biking', 'playground', 'sitting(again)', 'none',
                'break between activities']
    act_label = np.array(act_label)
    real_act_name = list()

    count = 0
    for i in range(12):
        if count == 12:
            break

        seen_flag = False
        for j in range(len(act_label)):
            if act_label[j] == i + 1:
                seen_flag = True
        if seen_flag == False:
            count += 1
            for k in range(len(act_label)):
                if act_label[k] > i + 1:
                    act_label[k] -= 1

        real_act_name.append(act_name[count])
        #print(act_name[count])
        count += 1
    return act_label,real_act_name

def act_boxplot(act_label, act_name,pred_y,test_y,model_name):
    #creates a boxplot for the error between predicted y-value and actual y-value for each activity type in the test set.
    # Also creates a folder and saves the figures.
    pred_error=pred_y-test_y
    df = pd.DataFrame(list(zip(pred_y, test_y, act_label,pred_error)),
                      columns=['predicted_y', 'actual_y', 'activity_label', 'prediction_error'])
    # Parent Directory path
    parent_dir = r'C:\Users\jcb\Desktop\speciale\models'
    min_y = pred_error.min()
    max_y = pred_error.max()
    path = os.path.join(parent_dir, model_name,"Boxplots")
    os.mkdir(path)
    for i in range(len(act_name)):
        plt.figure(i+1, figsize=(10, 10))
        df_act_label = df.loc[df['activity_label'] == i+1]
        plt.boxplot(df_act_label["prediction_error"])
        plt.title("boxplot for: "+act_name[i])
        plt.ylim(top=max_y+0.1)
        plt.ylim(bottom=min_y-0.1)
        plt.axhline(y=0, color='r', linestyle='--')
        figure_name =(path + '\\' +'activity'+ str(i+1)+'_'+act_name[i]+'_boxplot.png')
        plt.savefig(figure_name)
        plt.close()
        #plt.show()


def act_boxplot_profile(act_label, act_name,pred_y,test_y,model_name):
    #creates a boxplot for the error between predicted y-value and actual y-value for each activity type in the test set.
    # Also creates a folder and saves the figures.
    pred_error=pred_y-test_y
    df = pd.DataFrame(list(zip(pred_y, test_y, act_label,pred_error)),
                      columns=['predicted_y', 'actual_y', 'activity_label', 'prediction_error'])
    # Parent Directory path
    parent_dir = r'C:\Users\jcb\Desktop\speciale\models'
    min_y = pred_error.min()
    max_y = pred_error.max()
    path = os.path.join(parent_dir, model_name,"Boxplots","profile")
    os.mkdir(path)
    dict_act={}
    for i in range(len(act_name)):
        df_act_label = df.loc[df['activity_label'] == i+1]
        temp = df_act_label["prediction_error"]
        dict_act[act_name[i]]= temp
    fig, ax = plt.subplots()
    ax.boxplot(dict_act.values())
    ax.set_xticklabels(dict_act.keys())
    plt.xticks(rotation=90)
    plt.axhline(y=0, color='r', linestyle='--')
    fig.subplots_adjust(bottom=0.4)
    figure_name =(path + '\\' +'activity_profile_boxplot.png')
    plt.savefig(figure_name)
    plt.close()
    plt.show()
